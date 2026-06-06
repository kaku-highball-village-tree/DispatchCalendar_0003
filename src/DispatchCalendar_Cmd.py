#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""Dispatch calendar Excel-to-TSV converter (CMD version)."""

from __future__ import annotations

import sys
import traceback
import re
import json
import argparse
import ctypes
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError

SCOPES: list[str] = ["https://www.googleapis.com/auth/calendar.events"]
CREDENTIALS_FILE = Path("credentials") / "credentials.json"
TOKEN_FILE = Path("token") / "token.json"
TIME_ZONE = "Asia/Tokyo"
CALENDAR_ID = "primary"
GOOGLE_CALENDAR_ID_FILE = Path("google_calendar_id.txt")
DIALOG_TITLE = "DispatchCalendar DnD"


def get_google_calendar_id() -> str:
    """Load the target Google Calendar ID from google_calendar_id.txt, or use primary."""
    if not GOOGLE_CALENDAR_ID_FILE.exists():
        return CALENDAR_ID

    with GOOGLE_CALENDAR_ID_FILE.open(mode="r", encoding="utf-8") as obj_calendar_id_file:
        for psz_line in obj_calendar_id_file:
            psz_calendar_id: str = psz_line.strip()
            if psz_calendar_id != "":
                return psz_calendar_id

    return CALENDAR_ID


def write_error_text(psz_excel_file_path: str, psz_error_message: str) -> str:
    """Write an error text file beside the source Excel file."""
    obj_excel_path: Path = Path(psz_excel_file_path)
    psz_error_file_name: str = f"{obj_excel_path.stem}_error.txt"
    obj_error_file_path: Path = obj_excel_path.with_name(psz_error_file_name)

    with obj_error_file_path.open(mode="w", encoding="utf-8", newline="\r\n") as obj_error_file:
        obj_error_file.write(psz_error_message)

    return str(obj_error_file_path)


def show_auto_close_info_dialog(psz_message_text: str, i_timeout_milliseconds: int = 10000) -> None:
    """Show an OK information dialog that auto-closes after timeout."""
    try:
        ctypes.windll.user32.MessageBoxTimeoutW(
            0,
            psz_message_text,
            DIALOG_TITLE,
            0x00000040,  # MB_ICONINFORMATION
            0,
            i_timeout_milliseconds,
        )
    except Exception:
        return


def normalize_line_breaks_and_trim(psz_text: str) -> str:
    """Normalize line breaks and trim spaces for strict matching."""
    psz_normalized_text: str = psz_text.replace("\r\n", "\n").strip()
    return psz_normalized_text


def should_skip_row_by_first_column(list_source_row: list[str], list_skip_keywords: list[str]) -> bool:
    """Return True when first-column text matches any normalized skip keyword exactly."""
    if len(list_source_row) == 0:
        return False

    psz_first_column_text: str = normalize_line_breaks_and_trim(list_source_row[0])
    if psz_first_column_text == "":
        return False

    for psz_skip_keyword in list_skip_keywords:
        psz_normalized_skip_keyword: str = normalize_line_breaks_and_trim(psz_skip_keyword)
        if psz_first_column_text == psz_normalized_skip_keyword:
            return True

    return False




def is_shifted_continuation_row(list_previous_row: list[str], list_current_row: list[str]) -> bool:
    """Detect left-shifted continuation rows that should map A/B into C/D."""
    if len(list_previous_row) < 4 or len(list_current_row) < 2:
        return False

    psz_previous_name: str = normalize_line_breaks_and_trim(list_previous_row[0])
    psz_previous_car_number: str = normalize_line_breaks_and_trim(list_previous_row[1])
    psz_previous_spare_car: str = normalize_line_breaks_and_trim(list_previous_row[2])
    psz_current_first_value: str = normalize_line_breaks_and_trim(list_current_row[0])
    psz_current_second_value: str = normalize_line_breaks_and_trim(list_current_row[1])

    if psz_previous_name == "" or psz_previous_car_number == "" or psz_previous_spare_car == "":
        return False

    if psz_current_first_value == "" or psz_current_second_value == "":
        return False

    for psz_remaining_value in list_current_row[2:]:
        if normalize_line_breaks_and_trim(psz_remaining_value) != "":
            return False

    return True


def shift_row_to_the_right(list_current_row: list[str], iShiftCount: int) -> list[str]:
    """Return a row shifted right by adding empty cells at the beginning."""
    return [""] * iShiftCount + list_current_row
def merge_continuation_rows(list_source_rows: list[list[str]]) -> list[list[str]]:
    """Merge continuation rows into previous row by joining values with newlines per column."""
    list_merged_rows: list[list[str]] = []

    for list_current_row in list_source_rows:
        b_has_previous_row: bool = len(list_merged_rows) > 0
        b_is_blank_first_column_continuation: bool = b_has_previous_row and normalize_line_breaks_and_trim(list_current_row[0]) == ""

        b_has_shift_adjustment: bool = b_has_previous_row and is_shifted_continuation_row(list_merged_rows[-1], list_current_row)
        if b_has_shift_adjustment:
            list_adjusted_current_row: list[str] = shift_row_to_the_right(list_current_row, 2)
        else:
            list_adjusted_current_row = list_current_row

        b_is_continuation_row: bool = b_has_previous_row and (b_is_blank_first_column_continuation or b_has_shift_adjustment)

        if not b_is_continuation_row:
            list_merged_rows.append(list_current_row[:])
            continue

        list_previous_row: list[str] = list_merged_rows[-1]
        i_previous_column_count: int = len(list_previous_row)
        i_current_column_count: int = len(list_adjusted_current_row)
        i_max_column_count: int = i_previous_column_count if i_previous_column_count > i_current_column_count else i_current_column_count

        if i_previous_column_count < i_max_column_count:
            list_previous_row.extend([""] * (i_max_column_count - i_previous_column_count))

        for iColumnIndex in range(i_max_column_count):
            psz_current_value: str = ""
            if iColumnIndex < i_current_column_count:
                psz_current_value = list_adjusted_current_row[iColumnIndex]

            if normalize_line_breaks_and_trim(psz_current_value) == "":
                continue

            psz_previous_value: str = list_previous_row[iColumnIndex]
            if normalize_line_breaks_and_trim(psz_previous_value) == "":
                list_previous_row[iColumnIndex] = psz_current_value
            else:
                list_previous_row[iColumnIndex] = f"{psz_previous_value}\n{psz_current_value}"

    return list_merged_rows




def escape_newlines_in_cell_text(psz_cell_text: str) -> str:
    """Normalize embedded newlines in cell text."""
    psz_normalized_text: str = psz_cell_text.replace("\r\n", "\n").replace("\r", "\n")
    return psz_normalized_text


def expand_rows_by_embedded_newlines(
    list_source_rows: list[list[str]],
    i_fixed_column_count: int = 3,
) -> list[list[str]]:
    """Expand embedded newlines into physical TSV rows while blanking fixed leading columns on continuation lines."""
    list_expanded_rows: list[list[str]] = []

    for list_source_row in list_source_rows:
        list_split_cells: list[list[str]] = []
        i_max_split_count: int = 1
        b_has_multiline_non_fixed_column: bool = False

        for i_column_index, psz_cell_text in enumerate(list_source_row):
            list_cell_lines: list[str] = psz_cell_text.split("\n")
            list_split_cells.append(list_cell_lines)
            if len(list_cell_lines) > i_max_split_count:
                i_max_split_count = len(list_cell_lines)
            if i_column_index >= i_fixed_column_count and len(list_cell_lines) > 1:
                b_has_multiline_non_fixed_column = True

        for i_row_index in range(i_max_split_count):
            list_expanded_row: list[str] = []

            for i_column_index, list_cell_lines in enumerate(list_split_cells):
                psz_cell_value: str = list_cell_lines[i_row_index] if i_row_index < len(list_cell_lines) else ""

                b_should_blank_fixed_column: bool = (
                    i_row_index > 0
                    and b_has_multiline_non_fixed_column
                    and i_column_index < i_fixed_column_count
                    and len(list_cell_lines) == 1
                )
                if b_should_blank_fixed_column:
                    psz_cell_value = ""

                list_expanded_row.append(psz_cell_value)

            list_expanded_rows.append(list_expanded_row)

    return list_expanded_rows


def normalize_file_stem_for_step_output(psz_file_stem: str) -> str:
    """Replace half/full-width spaces in file stem with underscore."""
    return re.sub(r"[ \u3000]+", "_", psz_file_stem)


def create_step0001_tsv_from_tsv(psz_tsv_file_path: str) -> str:
    """Create step0001 TSV by removing spare-car column from the generated TSV."""
    obj_tsv_file_path: Path = Path(psz_tsv_file_path)
    psz_normalized_stem: str = normalize_file_stem_for_step_output(obj_tsv_file_path.stem)
    obj_step_tsv_path: Path = obj_tsv_file_path.with_name(f"{psz_normalized_stem}_step0001.tsv")

    list_output_lines: list[str] = []
    with obj_tsv_file_path.open(mode="r", encoding="utf-8", newline="") as obj_input_file:
        for psz_line in obj_input_file:
            psz_line_without_newline: str = psz_line.rstrip("\r\n")
            list_columns: list[str] = psz_line_without_newline.split("\t")
            if len(list_columns) >= 3:
                del list_columns[2]
            list_output_lines.append("\t".join(list_columns))

    with obj_step_tsv_path.open(mode="w", encoding="utf-8", newline="\r\n") as obj_output_file:
        for psz_output_line in list_output_lines:
            obj_output_file.write(psz_output_line + "\n")

    return str(obj_step_tsv_path)



def create_step0001_5_tsv_from_step0001_tsv(psz_step0001_tsv_path: str) -> str:
    """Create step0001_5 TSV by expanding horizontal work slots into vertical work sets."""
    obj_step0001_path: Path = Path(psz_step0001_tsv_path)
    psz_step_stem: str = obj_step0001_path.stem
    psz_output_stem: str = psz_step_stem[:-9] if psz_step_stem.endswith("_step0001") else psz_step_stem
    obj_step0001_5_path: Path = obj_step0001_path.with_name(f"{psz_output_stem}_step0001_5.tsv")

    with obj_step0001_path.open(mode="r", encoding="utf-8", newline="") as obj_input_file:
        list_lines: list[str] = [psz_line.rstrip("\r\n") for psz_line in obj_input_file]

    if len(list_lines) <= 2:
        with obj_step0001_5_path.open(mode="w", encoding="utf-8", newline="\r\n") as obj_output_file:
            for psz_line in list_lines:
                obj_output_file.write(psz_line + "\n")
        return str(obj_step0001_5_path)

    list_output_lines: list[str] = list_lines[:2]
    list_current_group_rows: list[list[str]] = []
    psz_current_name: str = ""
    psz_current_car_no: str = ""
    psz_last_car_no: str = ""

    def normalize_step0001_5_row(psz_line: str) -> list[str]:
        list_columns: list[str] = psz_line.split("\t")
        if len(list_columns) < 8:
            list_columns.extend([""] * (8 - len(list_columns)))
        return list_columns

    def flush_current_group() -> None:
        nonlocal list_current_group_rows, psz_current_name, psz_current_car_no
        if len(list_current_group_rows) == 0:
            return

        for i_slot_index in range(2, 8):
            list_slot_values: list[str] = []
            for list_group_row in list_current_group_rows:
                psz_slot_value: str = ""
                if i_slot_index < len(list_group_row):
                    psz_slot_value = list_group_row[i_slot_index]
                list_slot_values.append(psz_slot_value)

            b_is_empty_slot: bool = all(normalize_line_breaks_and_trim(psz_value) == "" for psz_value in list_slot_values)
            if b_is_empty_slot:
                break

            list_output_lines.append("\t".join([psz_current_name, psz_current_car_no, list_slot_values[0]]))
            for psz_continuation_value in list_slot_values[1:]:
                if normalize_line_breaks_and_trim(psz_continuation_value) == "":
                    continue
                list_output_lines.append("\t".join(["", "", psz_continuation_value]))

        list_current_group_rows = []
        psz_current_name = ""
        psz_current_car_no = ""

    for psz_line in list_lines[2:]:
        list_columns = normalize_step0001_5_row(psz_line)
        psz_name: str = list_columns[0].strip()
        psz_car_no: str = list_columns[1].strip()
        list_slot_values: list[str] = [list_columns[i].strip() for i in range(2, 8)]
        b_is_blank_row: bool = psz_name == "" and psz_car_no == "" and all(psz_slot_value == "" for psz_slot_value in list_slot_values)
        if b_is_blank_row:
            continue

        if psz_name != "":
            flush_current_group()
            psz_current_name = psz_name
            psz_current_car_no = psz_car_no if psz_car_no != "" else psz_last_car_no
            if psz_current_car_no != "":
                psz_last_car_no = psz_current_car_no
            list_current_group_rows.append(list_columns)
            continue

        if len(list_current_group_rows) == 0:
            continue

        if psz_car_no != "":
            psz_current_car_no = psz_car_no
            psz_last_car_no = psz_car_no
        list_current_group_rows.append(list_columns)

    flush_current_group()

    with obj_step0001_5_path.open(mode="w", encoding="utf-8", newline="\r\n") as obj_output_file:
        for psz_output_line in list_output_lines:
            obj_output_file.write(psz_output_line + "\n")

    return str(obj_step0001_5_path)

def parse_step0001_tsv_to_calendar_records(psz_step0001_tsv_path: str) -> list[dict[str, Any]]:
    """Parse step0001 TSV into per-person calendar records."""
    obj_step0001_tsv_path: Path = Path(psz_step0001_tsv_path)
    with obj_step0001_tsv_path.open(mode="r", encoding="utf-8", newline="") as obj_input_file:
        list_lines: list[str] = [psz_line.rstrip("\r\n") for psz_line in obj_input_file]

    if len(list_lines) < 3:
        return []

    psz_work_date_text: str = list_lines[0]
    list_data_lines: list[str] = list_lines[2:]

    list_records: list[dict[str, Any]] = []
    obj_current_record: dict[str, Any] | None = None

    for psz_line in list_data_lines:
        list_columns: list[str] = psz_line.split("\t")
        if len(list_columns) < 8:
            list_columns.extend([""] * (8 - len(list_columns)))

        psz_name: str = list_columns[0].strip()
        psz_car_no: str = list_columns[1].strip()
        list_slot_values: list[str] = [list_columns[i].strip() for i in range(2, 8)]

        b_is_blank_row: bool = psz_name == "" and psz_car_no == "" and all(psz_slot == "" for psz_slot in list_slot_values)
        if b_is_blank_row:
            continue

        if psz_name != "":
            if obj_current_record is not None:
                list_records.append(obj_current_record)
            obj_current_record = {
                "name": psz_name,
                "car_nos": [],
                "slots": {str(i): "" for i in range(1, 7)},
                "work_date_text": psz_work_date_text,
            }

        if obj_current_record is None:
            continue

        if psz_car_no != "":
            list_car_nos: list[str] = obj_current_record["car_nos"]
            if psz_car_no not in list_car_nos and len(list_car_nos) < 3:
                list_car_nos.append(psz_car_no)

        for i_slot_index, psz_slot_value in enumerate(list_slot_values, start=1):
            if psz_slot_value == "":
                continue
            psz_slot_key: str = str(i_slot_index)
            psz_existing: str = obj_current_record["slots"][psz_slot_key]
            if psz_existing == "":
                obj_current_record["slots"][psz_slot_key] = psz_slot_value
            else:
                obj_current_record["slots"][psz_slot_key] = f"{psz_existing}\n{psz_slot_value}"

    if obj_current_record is not None:
        list_records.append(obj_current_record)

    for obj_record in list_records:
        list_car_nos = obj_record["car_nos"]
        obj_record["car_no"] = ",".join(list_car_nos)
        obj_record["car_no_display"] = "/".join(list_car_nos)
        psz_car_for_title: str = obj_record["car_no_display"] if obj_record["car_no_display"] != "" else "車番未設定"
        list_slot_lines: list[list[str]] = [obj_record["slots"][str(i_slot_index)].split("\n") for i_slot_index in range(1, 7)]
        i_max_slot_line_count: int = 1
        for list_slot_line in list_slot_lines:
            if len(list_slot_line) > i_max_slot_line_count:
                i_max_slot_line_count = len(list_slot_line)

        def build_slot_line(i_line_index: int) -> str:
            list_values: list[str] = [
                list_slot_line[i_line_index] if i_line_index < len(list_slot_line) else "" for list_slot_line in list_slot_lines
            ]
            list_non_blank_values: list[str] = [psz_value for psz_value in list_values if psz_value != ""]
            return " ".join(list_non_blank_values)

        psz_first_slot_line: str = build_slot_line(0)
        list_title_parts: list[str] = [obj_record["name"], psz_car_for_title]
        if psz_first_slot_line != "":
            list_title_parts.append(psz_first_slot_line)
        obj_record["title_text"] = " ".join(list_title_parts)

        list_body_lines: list[str] = []
        if i_max_slot_line_count >= 2:
            psz_second_slot_line: str = build_slot_line(1)
            if psz_second_slot_line != "":
                list_body_lines.append(psz_second_slot_line)
        else:
            if psz_first_slot_line != "":
                list_body_lines.append(psz_first_slot_line)

        for i_line_index in range(2, i_max_slot_line_count):
            psz_extra_slot_line: str = build_slot_line(i_line_index)
            if psz_extra_slot_line != "":
                list_body_lines.append(psz_extra_slot_line)

        obj_record["body_text"] = "\n".join(list_body_lines)

    return list_records


def create_step0002_outputs_from_step0001_tsv(psz_step0001_tsv_path: str) -> tuple[str, str]:
    """Create step0002 TSV and JSON (NDJSON) from step0001 or step0001_5 TSV."""
    obj_step0001_tsv_path: Path = Path(psz_step0001_tsv_path)
    psz_step_stem: str = obj_step0001_tsv_path.stem
    if psz_step_stem.endswith("_step0001_5"):
        psz_output_stem: str = psz_step_stem[:-11]
    elif psz_step_stem.endswith("_step0001"):
        psz_output_stem = psz_step_stem[:-9]
    else:
        psz_output_stem = psz_step_stem

    obj_step0002_tsv_path: Path = obj_step0001_tsv_path.with_name(f"{psz_output_stem}_step0002.tsv")
    obj_step0002_json_path: Path = obj_step0001_tsv_path.with_name(f"{psz_output_stem}_step0002.json")

    list_records: list[dict[str, Any]] = parse_step0001_tsv_to_calendar_records(str(obj_step0001_tsv_path))

    list_tsv_columns: list[str] = [
        "name",
        "car_no_display",
        "car_no",
        "car_nos_joined",
        "slot1",
        "slot2",
        "slot3",
        "slot4",
        "slot5",
        "slot6",
        "title_text",
        "body_text",
        "work_date_text",
    ]
    with obj_step0002_tsv_path.open(mode="w", encoding="utf-8", newline="\r\n") as obj_tsv_file:
        obj_tsv_file.write("\t".join(list_tsv_columns) + "\n")
        for obj_record in list_records:
            list_slot_lines: list[list[str]] = [obj_record["slots"][str(i_slot_index)].split("\n") for i_slot_index in range(1, 7)]
            i_max_slot_line_count: int = 1
            for list_slot_line in list_slot_lines:
                if len(list_slot_line) > i_max_slot_line_count:
                    i_max_slot_line_count = len(list_slot_line)

            for i_line_index in range(i_max_slot_line_count):
                list_slot_values: list[str] = [
                    list_slot_line[i_line_index] if i_line_index < len(list_slot_line) else "" for list_slot_line in list_slot_lines
                ]
                list_row_values: list[str] = [
                    obj_record["name"] if i_line_index == 0 else "",
                    obj_record["car_no_display"] if i_line_index == 0 else "",
                    obj_record["car_no"] if i_line_index == 0 else "",
                    obj_record["car_no"] if i_line_index == 0 else "",
                    list_slot_values[0],
                    list_slot_values[1],
                    list_slot_values[2],
                    list_slot_values[3],
                    list_slot_values[4],
                    list_slot_values[5],
                    obj_record["title_text"] if i_line_index == 0 else "",
                    obj_record["body_text"].replace("\n", "\\n") if i_line_index == 0 else "",
                    obj_record["work_date_text"] if i_line_index == 0 else "",
                ]
                obj_tsv_file.write("\t".join(list_row_values) + "\n")

    with obj_step0002_json_path.open(mode="w", encoding="utf-8", newline="\n") as obj_json_file:
        for obj_record in list_records:
            obj_json_file.write(json.dumps(obj_record, ensure_ascii=False) + "\n")

    return str(obj_step0002_tsv_path), str(obj_step0002_json_path)




def create_step0003_tsv_from_step0002_tsv(psz_step0002_tsv_path: str) -> str:
    """Create step0003 TSV by removing columns 1-10 from step0002 TSV."""
    obj_step0002_path: Path = Path(psz_step0002_tsv_path)
    psz_output_stem: str = obj_step0002_path.stem[:-9] if obj_step0002_path.stem.endswith("_step0002") else obj_step0002_path.stem
    obj_step0003_path: Path = obj_step0002_path.with_name(f"{psz_output_stem}_step0003.tsv")

    with obj_step0002_path.open(mode="r", encoding="utf-8", newline="") as obj_input_file:
        list_lines: list[str] = [psz_line.rstrip("\r\n") for psz_line in obj_input_file]

    list_output_lines: list[str] = []
    for psz_line in list_lines:
        list_columns: list[str] = psz_line.split("\t")
        if len(list_columns) < 10:
            raise ValueError("step0002 TSV row must have at least 10 columns")
        list_output_lines.append("\t".join(list_columns[10:]))

    with obj_step0003_path.open(mode="w", encoding="utf-8", newline="\r\n") as obj_output_file:
        for psz_output_line in list_output_lines:
            obj_output_file.write(psz_output_line + "\n")

    return str(obj_step0003_path)

def convert_japanese_era_date_text_to_iso(psz_work_date_text: str) -> str:
    """Convert Japanese era style date text like '令和8年 5月 5日（火）' to ISO date (YYYY-MM-DD)."""
    psz_normalized_text: str = psz_work_date_text
    psz_normalized_text = psz_normalized_text.translate(str.maketrans("０１２３４５６７８９", "0123456789"))
    psz_normalized_text = re.sub(r"[\u3000\s]+", " ", psz_normalized_text).strip()
    psz_normalized_text = re.sub(r"[（(][^)）]*[）)]", "", psz_normalized_text).strip()

    obj_match: re.Match[str] | None = re.search(r"令和\s*(\d+)\s*年\s*(\d+)\s*月\s*(\d+)\s*日", psz_normalized_text)
    if obj_match is not None:
        i_reiwa_year: int = int(obj_match.group(1))
        i_year: int = 2018 + i_reiwa_year
        i_month: int = int(obj_match.group(2))
        i_day: int = int(obj_match.group(3))
        return f"{i_year:04d}-{i_month:02d}-{i_day:02d}"

    obj_western_match: re.Match[str] | None = re.search(r"(\d{4})\s*年\s*(\d+)\s*月\s*(\d+)\s*日", psz_normalized_text)
    if obj_western_match is not None:
        i_year = int(obj_western_match.group(1))
        i_month = int(obj_western_match.group(2))
        i_day = int(obj_western_match.group(3))
        return f"{i_year:04d}-{i_month:02d}-{i_day:02d}"

    raise ValueError(f"Unable to parse work_date_text: {psz_work_date_text}")


def create_step0004_tsv_from_step0003_tsv(psz_step0003_tsv_path: str) -> str:
    """Create step0004 TSV by inserting work_date_iso column derived from work_date_text."""
    obj_step0003_path: Path = Path(psz_step0003_tsv_path)
    psz_output_stem: str = obj_step0003_path.stem[:-9] if obj_step0003_path.stem.endswith("_step0003") else obj_step0003_path.stem
    obj_step0004_path: Path = obj_step0003_path.with_name(f"{psz_output_stem}_step0004.tsv")

    with obj_step0003_path.open(mode="r", encoding="utf-8", newline="") as obj_input_file:
        list_lines: list[str] = [psz_line.rstrip("\r\n") for psz_line in obj_input_file]

    if len(list_lines) == 0:
        with obj_step0004_path.open(mode="w", encoding="utf-8", newline="\r\n") as obj_output_file:
            obj_output_file.write("")
        return str(obj_step0004_path)

    list_header_columns: list[str] = list_lines[0].split("\t")
    if "work_date_text" not in list_header_columns:
        raise ValueError("step0003 TSV must include work_date_text column")

    i_work_date_text_index: int = list_header_columns.index("work_date_text")
    i_insert_index: int = i_work_date_text_index + 1
    list_output_lines: list[str] = []

    list_new_header_columns: list[str] = list_header_columns[:]
    list_new_header_columns.insert(i_insert_index, "work_date_iso")
    list_output_lines.append("\t".join(list_new_header_columns))

    for psz_data_line in list_lines[1:]:
        list_columns: list[str] = psz_data_line.split("\t")
        if len(list_columns) < len(list_header_columns):
            list_columns.extend([""] * (len(list_header_columns) - len(list_columns)))

        psz_work_date_text: str = list_columns[i_work_date_text_index]
        psz_work_date_iso: str = ""
        if psz_work_date_text.strip() != "":
            psz_work_date_iso = convert_japanese_era_date_text_to_iso(psz_work_date_text)

        list_columns.insert(i_insert_index, psz_work_date_iso)
        list_output_lines.append("\t".join(list_columns))

    with obj_step0004_path.open(mode="w", encoding="utf-8", newline="\r\n") as obj_output_file:
        for psz_output_line in list_output_lines:
            obj_output_file.write(psz_output_line + "\n")

    return str(obj_step0004_path)


def create_step0005_tsv_from_step0004_tsv(psz_step0004_tsv_path: str) -> str:
    """Create step0005 TSV by trimming trailing tab-only empty cells from data rows."""
    obj_step0004_path: Path = Path(psz_step0004_tsv_path)
    psz_output_stem: str = obj_step0004_path.stem[:-9] if obj_step0004_path.stem.endswith("_step0004") else obj_step0004_path.stem
    obj_step0005_path: Path = obj_step0004_path.with_name(f"{psz_output_stem}_step0005.tsv")

    with obj_step0004_path.open(mode="r", encoding="utf-8", newline="") as obj_input_file:
        list_lines: list[str] = [psz_line.rstrip("\r\n") for psz_line in obj_input_file]

    if len(list_lines) == 0:
        with obj_step0005_path.open(mode="w", encoding="utf-8", newline="\r\n") as obj_output_file:
            obj_output_file.write("")
        return str(obj_step0005_path)

    list_output_lines: list[str] = [list_lines[0]]
    for psz_data_line in list_lines[1:]:
        list_output_lines.append(psz_data_line.rstrip("	"))

    with obj_step0005_path.open(mode="w", encoding="utf-8", newline="\r\n") as obj_output_file:
        for psz_output_line in list_output_lines:
            obj_output_file.write(psz_output_line + "\n")

    return str(obj_step0005_path)


def create_step0006_tsv_from_step0005_tsv(psz_step0005_tsv_path: str) -> str:
    """Create step0006 TSV by removing work_date_text column."""
    obj_step0005_path: Path = Path(psz_step0005_tsv_path)
    psz_output_stem: str = obj_step0005_path.stem[:-9] if obj_step0005_path.stem.endswith("_step0005") else obj_step0005_path.stem
    obj_step0006_path: Path = obj_step0005_path.with_name(f"{psz_output_stem}_step0006.tsv")

    with obj_step0005_path.open(mode="r", encoding="utf-8", newline="") as obj_input_file:
        list_lines: list[str] = [psz_line.rstrip("\r\n") for psz_line in obj_input_file]

    if len(list_lines) == 0:
        with obj_step0006_path.open(mode="w", encoding="utf-8", newline="\r\n") as obj_output_file:
            obj_output_file.write("")
        return str(obj_step0006_path)

    list_header_columns: list[str] = list_lines[0].split("\t")
    if "work_date_text" not in list_header_columns:
        raise ValueError("step0005 TSV must include work_date_text column")

    i_work_date_text_index: int = list_header_columns.index("work_date_text")
    list_output_lines: list[str] = []

    for psz_line in list_lines:
        list_columns: list[str] = psz_line.split("\t")
        if i_work_date_text_index < len(list_columns):
            del list_columns[i_work_date_text_index]
        list_output_lines.append("\t".join(list_columns))

    with obj_step0006_path.open(mode="w", encoding="utf-8", newline="\r\n") as obj_output_file:
        for psz_output_line in list_output_lines:
            obj_output_file.write(psz_output_line + "\n")

    return str(obj_step0006_path)


def create_step0007_tsv_from_step0006_tsv(psz_step0006_tsv_path: str) -> str:
    """Create step0007 TSV by removing blank lines from step0006 TSV."""
    obj_step0006_path: Path = Path(psz_step0006_tsv_path)
    psz_output_stem: str = obj_step0006_path.stem[:-9] if obj_step0006_path.stem.endswith("_step0006") else obj_step0006_path.stem
    obj_step0007_path: Path = obj_step0006_path.with_name(f"{psz_output_stem}_step0007.tsv")

    with obj_step0006_path.open(mode="r", encoding="utf-8", newline="") as obj_input_file:
        list_lines: list[str] = [psz_line.rstrip("\r\n") for psz_line in obj_input_file]

    list_output_lines: list[str] = []
    for psz_line in list_lines:
        if psz_line.strip() == "":
            continue
        list_output_lines.append(psz_line)

    with obj_step0007_path.open(mode="w", encoding="utf-8", newline="\r\n") as obj_output_file:
        for psz_output_line in list_output_lines:
            obj_output_file.write(psz_output_line + "\n")

    return str(obj_step0007_path)


def get_google_credentials() -> Credentials:
    """Load credentials from token.json or run OAuth flow if needed."""
    if not CREDENTIALS_FILE.exists():
        raise FileNotFoundError("credentials/credentials.json が見つかりません。")

    obj_credentials: Credentials | None = None
    if TOKEN_FILE.exists():
        try:
            obj_credentials = Credentials.from_authorized_user_file(str(TOKEN_FILE), SCOPES)
        except Exception:
            obj_credentials = None

    if not obj_credentials or not obj_credentials.valid:
        if obj_credentials and obj_credentials.expired and obj_credentials.refresh_token:
            try:
                obj_credentials.refresh(Request())
            except Exception:
                obj_credentials = None

        if not obj_credentials or not obj_credentials.valid:
            obj_flow = InstalledAppFlow.from_client_secrets_file(str(CREDENTIALS_FILE), SCOPES)
            obj_credentials = obj_flow.run_local_server(port=0)

        TOKEN_FILE.parent.mkdir(parents=True, exist_ok=True)
        TOKEN_FILE.write_text(obj_credentials.to_json(), encoding="utf-8")

    return obj_credentials


def create_google_calendar_events_from_step0007_tsv(psz_step0007_tsv_path: str) -> tuple[int, int]:
    """Register events from step0007 TSV; skip invalid rows and write *_error.txt."""
    obj_step0007_path: Path = Path(psz_step0007_tsv_path)
    with obj_step0007_path.open(mode="r", encoding="utf-8", newline="") as obj_input_file:
        list_lines: list[str] = [psz_line.rstrip("\r\n") for psz_line in obj_input_file]

    if len(list_lines) == 0:
        return 0, 0

    list_header_columns: list[str] = list_lines[0].split("\t")
    for psz_required_column in ["title_text", "body_text", "work_date_iso"]:
        if psz_required_column not in list_header_columns:
            raise ValueError(f"step0007 TSV must include {psz_required_column} column")

    i_title_index: int = list_header_columns.index("title_text")
    i_body_index: int = list_header_columns.index("body_text")
    i_work_date_iso_index: int = list_header_columns.index("work_date_iso")

    obj_service = build("calendar", "v3", credentials=get_google_credentials())
    psz_calendar_id: str = get_google_calendar_id()

    i_success_count: int = 0
    i_skip_count: int = 0
    list_skip_messages: list[str] = []

    for i_line_number, psz_line in enumerate(list_lines[1:], start=2):
        list_columns: list[str] = psz_line.split("\t")
        if len(list_columns) <= max(i_title_index, i_body_index, i_work_date_iso_index):
            i_skip_count += 1
            list_skip_messages.append(f"line={i_line_number}, reason=required columns missing in row")
            continue

        psz_title: str = list_columns[i_title_index].strip()
        psz_body: str = list_columns[i_body_index].replace("\\n", "\n").strip()
        psz_work_date_iso: str = list_columns[i_work_date_iso_index].strip()

        if psz_title == "":
            i_skip_count += 1
            list_skip_messages.append(
                f"line={i_line_number}, reason=title_text is empty, work_date_iso={psz_work_date_iso}"
            )
            continue

        if psz_work_date_iso == "":
            i_skip_count += 1
            list_skip_messages.append(f"line={i_line_number}, reason=work_date_iso is empty, title_text={psz_title}")
            continue

        try:
            if "T" in psz_work_date_iso:
                obj_start_datetime: datetime = datetime.fromisoformat(psz_work_date_iso)
                obj_end_datetime: datetime = obj_start_datetime + timedelta(hours=1)
                obj_event_body: dict[str, object] = {
                    "summary": psz_title,
                    "location": "",
                    "description": psz_body,
                    "start": {"dateTime": obj_start_datetime.isoformat(), "timeZone": TIME_ZONE},
                    "end": {"dateTime": obj_end_datetime.isoformat(), "timeZone": TIME_ZONE},
                }
            else:
                obj_start_date: datetime = datetime.strptime(psz_work_date_iso, "%Y-%m-%d")
                obj_end_date: datetime = obj_start_date + timedelta(days=1)
                obj_event_body = {
                    "summary": psz_title,
                    "location": "",
                    "description": psz_body,
                    "start": {"date": obj_start_date.strftime("%Y-%m-%d"), "timeZone": TIME_ZONE},
                    "end": {"date": obj_end_date.strftime("%Y-%m-%d"), "timeZone": TIME_ZONE},
                }

            created_event = (
                obj_service.events()
                .insert(calendarId=psz_calendar_id, body=obj_event_body)
                .execute()
            )
            print(created_event.get("htmlLink", ""))
            i_success_count += 1
        except HttpError as obj_exception:
            i_skip_count += 1
            list_skip_messages.append(
                f"line={i_line_number}, reason={obj_exception}, work_date_iso={psz_work_date_iso}, title_text={psz_title}"
            )
        except Exception as obj_exception:
            i_skip_count += 1
            list_skip_messages.append(
                f"line={i_line_number}, reason={obj_exception}, work_date_iso={psz_work_date_iso}, title_text={psz_title}"
            )

    if len(list_skip_messages) > 0:
        write_error_text(str(obj_step0007_path), "\n".join(list_skip_messages))

    return i_success_count, i_skip_count


def delete_google_calendar_events_from_step0007_tsv(psz_step0007_tsv_path: str) -> tuple[int, int]:
    """Delete matching events from step0007 TSV; skip invalid rows and write *_error.txt."""
    obj_step0007_path: Path = Path(psz_step0007_tsv_path)
    with obj_step0007_path.open(mode="r", encoding="utf-8", newline="") as obj_input_file:
        list_lines: list[str] = [psz_line.rstrip("\r\n") for psz_line in obj_input_file]

    if len(list_lines) == 0:
        return 0, 0

    list_header_columns: list[str] = list_lines[0].split("\t")
    for psz_required_column in ["title_text", "body_text", "work_date_iso"]:
        if psz_required_column not in list_header_columns:
            raise ValueError(f"step0007 TSV must include {psz_required_column} column")

    i_title_index: int = list_header_columns.index("title_text")
    i_body_index: int = list_header_columns.index("body_text")
    i_work_date_iso_index: int = list_header_columns.index("work_date_iso")

    obj_service = build("calendar", "v3", credentials=get_google_credentials())
    psz_calendar_id: str = get_google_calendar_id()

    i_deleted_count: int = 0
    i_skip_count: int = 0
    list_skip_messages: list[str] = []

    for i_line_number, psz_line in enumerate(list_lines[1:], start=2):
        list_columns: list[str] = psz_line.split("\t")
        if len(list_columns) <= max(i_title_index, i_body_index, i_work_date_iso_index):
            i_skip_count += 1
            list_skip_messages.append(f"line={i_line_number}, reason=required columns missing in row")
            continue

        psz_title: str = list_columns[i_title_index].strip()
        psz_body: str = list_columns[i_body_index].replace("\\n", "\n").strip()
        psz_work_date_iso: str = list_columns[i_work_date_iso_index].strip()

        if psz_title == "" or psz_work_date_iso == "":
            i_skip_count += 1
            list_skip_messages.append(
                f"line={i_line_number}, reason=title_text/work_date_iso is empty, work_date_iso={psz_work_date_iso}"
            )
            continue

        try:
            if "T" in psz_work_date_iso:
                obj_start_datetime: datetime = datetime.fromisoformat(psz_work_date_iso)
                obj_time_min: datetime = obj_start_datetime.replace(hour=0, minute=0, second=0, microsecond=0)
            else:
                obj_time_min = datetime.strptime(psz_work_date_iso, "%Y-%m-%d")
            obj_time_max: datetime = obj_time_min + timedelta(days=1)

            obj_response = (
                obj_service.events()
                .list(
                    calendarId=psz_calendar_id,
                    timeMin=obj_time_min.isoformat() + "+09:00",
                    timeMax=obj_time_max.isoformat() + "+09:00",
                    singleEvents=True,
                )
                .execute()
            )

            list_items: list[dict[str, Any]] = obj_response.get("items", [])
            for obj_item in list_items:
                psz_summary: str = str(obj_item.get("summary", "")).strip()
                psz_description: str = str(obj_item.get("description", "")).strip()
                if psz_summary == psz_title and psz_description == psz_body:
                    psz_event_id: str = str(obj_item.get("id", ""))
                    if psz_event_id == "":
                        continue
                    obj_service.events().delete(calendarId=psz_calendar_id, eventId=psz_event_id).execute()
                    i_deleted_count += 1
        except HttpError as obj_exception:
            i_skip_count += 1
            list_skip_messages.append(
                f"line={i_line_number}, reason={obj_exception}, work_date_iso={psz_work_date_iso}, title_text={psz_title}"
            )
        except Exception as obj_exception:
            i_skip_count += 1
            list_skip_messages.append(
                f"line={i_line_number}, reason={obj_exception}, work_date_iso={psz_work_date_iso}, title_text={psz_title}"
            )

    if len(list_skip_messages) > 0:
        write_error_text(str(obj_step0007_path), "\n".join(list_skip_messages))

    return i_deleted_count, i_skip_count

def convert_excel_to_tsv(psz_excel_file_path: str) -> str:
    """Convert active sheet of an Excel file to UTF-8 TSV with CRLF line endings."""
    obj_excel_path: Path = Path(psz_excel_file_path)
    psz_tsv_file_name: str = f"{obj_excel_path.stem}.tsv"
    obj_tsv_file_path: Path = obj_excel_path.with_name(psz_tsv_file_name)

    obj_workbook: Any = load_workbook(filename=str(obj_excel_path), data_only=True)
    obj_active_sheet: Any = obj_workbook.active

    list_skip_keywords: list[str] = ["始業前点検"]

    list_source_rows: list[list[str]] = []
    for obj_row in obj_active_sheet.iter_rows(values_only=True):
        list_row_values: list[str] = []
        for obj_cell_value in obj_row:
            if obj_cell_value is None:
                psz_cell_text: str = ""
            else:
                psz_cell_text = escape_newlines_in_cell_text(str(obj_cell_value))
            list_row_values.append(psz_cell_text)

        if should_skip_row_by_first_column(list_row_values, list_skip_keywords):
            continue

        list_source_rows.append(list_row_values)

    list_normalized_rows: list[list[str]] = merge_continuation_rows(list_source_rows)

    list_expanded_rows: list[list[str]] = expand_rows_by_embedded_newlines(list_normalized_rows)

    with obj_tsv_file_path.open(mode="w", encoding="utf-8", newline="\r\n") as obj_tsv_file:
        for list_normalized_row in list_expanded_rows:
            psz_tsv_line: str = "\t".join(list_normalized_row)
            obj_tsv_file.write(psz_tsv_line + "\n")

    obj_workbook.close()
    return str(obj_tsv_file_path)


def main() -> int:
    list_arguments: list[str] = sys.argv
    obj_parser = argparse.ArgumentParser(add_help=False)
    obj_parser.add_argument("--mode", choices=["create", "delete"], default="create")
    obj_parser.add_argument("excel_file_paths", nargs="*")
    obj_parsed = obj_parser.parse_args(list_arguments[1:])

    if len(obj_parsed.excel_file_paths) < 1:
        psz_usage_message: str = "Usage: python DispatchCalendar_Cmd.py <excel_file_path1> [excel_file_path2 ...]"
        print(psz_usage_message)
        return 1

    list_excel_file_paths: list[str] = obj_parsed.excel_file_paths
    i_success_count: int = 0
    i_failure_count: int = 0

    for psz_excel_file_path in list_excel_file_paths:
        obj_excel_path: Path = Path(psz_excel_file_path)

        if not obj_excel_path.exists() or not obj_excel_path.is_file():
            psz_error_message: str = f"Input file not found: {psz_excel_file_path}"
            print(psz_error_message)
            write_error_text(psz_excel_file_path, psz_error_message)
            i_failure_count += 1
            continue

        if obj_excel_path.suffix.lower() != ".xlsx":
            psz_error_message = f"Only .xlsx files are supported: {psz_excel_file_path}"
            print(psz_error_message)
            write_error_text(psz_excel_file_path, psz_error_message)
            i_failure_count += 1
            continue

        try:
            psz_created_tsv_path: str = convert_excel_to_tsv(psz_excel_file_path)
            print(f"TSV created: {psz_created_tsv_path}")
            psz_created_step_tsv_path: str = create_step0001_tsv_from_tsv(psz_created_tsv_path)
            print(f"Step TSV created: {psz_created_step_tsv_path}")
            psz_step0001_5_tsv_path: str = create_step0001_5_tsv_from_step0001_tsv(psz_created_step_tsv_path)
            print(f"Step0001.5 TSV created: {psz_step0001_5_tsv_path}")
            psz_step0002_tsv_path, psz_step0002_json_path = create_step0002_outputs_from_step0001_tsv(psz_step0001_5_tsv_path)
            print(f"Step0002 TSV created: {psz_step0002_tsv_path}")
            print(f"Step0002 JSON created: {psz_step0002_json_path}")
            psz_step0003_tsv_path: str = create_step0003_tsv_from_step0002_tsv(psz_step0002_tsv_path)
            print(f"Step0003 TSV created: {psz_step0003_tsv_path}")
            psz_step0004_tsv_path: str = create_step0004_tsv_from_step0003_tsv(psz_step0003_tsv_path)
            print(f"Step0004 TSV created: {psz_step0004_tsv_path}")
            psz_step0005_tsv_path: str = create_step0005_tsv_from_step0004_tsv(psz_step0004_tsv_path)
            print(f"Step0005 TSV created: {psz_step0005_tsv_path}")
            psz_step0006_tsv_path: str = create_step0006_tsv_from_step0005_tsv(psz_step0005_tsv_path)
            print(f"Step0006 TSV created: {psz_step0006_tsv_path}")
            psz_step0007_tsv_path: str = create_step0007_tsv_from_step0006_tsv(psz_step0006_tsv_path)
            print(f"Step0007 TSV created: {psz_step0007_tsv_path}")
            if obj_parsed.mode == "delete":
                i_deleted_count, i_skipped_count = delete_google_calendar_events_from_step0007_tsv(psz_step0007_tsv_path)
                print(f"Google Calendar events deleted: {i_deleted_count}, skipped: {i_skipped_count}")
            else:
                show_auto_close_info_dialog("step0007.tsv を作成しました。")
                show_auto_close_info_dialog("カレンダーへの登録を開始します。", 5000)
                i_registered_count, i_skipped_count = create_google_calendar_events_from_step0007_tsv(psz_step0007_tsv_path)
                print(f"Google Calendar events created: {i_registered_count}, skipped: {i_skipped_count}")

            i_success_count += 1
        except Exception as obj_exception:  # noqa: BLE001
            psz_traceback_text: str = traceback.format_exc()
            psz_error_message = (
                "Failed to convert Excel to TSV.\n"
                f"Input: {psz_excel_file_path}\n"
                f"Error: {obj_exception}\n\n"
                f"Traceback:\n{psz_traceback_text}"
            )
            print(psz_error_message)
            write_error_text(psz_excel_file_path, psz_error_message)
            i_failure_count += 1

    print(f"Summary: success={i_success_count}, failure={i_failure_count}")
    if i_failure_count == 0:
        return 0
    return 1


if __name__ == "__main__":
    raise SystemExit(main())
